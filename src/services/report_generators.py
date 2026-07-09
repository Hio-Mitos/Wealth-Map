"""
WealthMap – Report Generators
One generator class per module + a master "full financial report".
Each class produces a PDF and optionally an XLSX.
"""

import os
import csv
import io
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Tuple

from src.services.report_engine import (
    ReportBuilder, STYLES, C_ACCENT, C_GREEN, C_RED, C_GOLD, C_MUTED,
    C_TEXT, C_CARD, C_BG, C_BORDER, _money_cell, _pct_cell, CHART_PALETTE
)
from src.models.database import (
    TransactionType, TransactionStatus, AccountType, AssetType,
    Transaction, PortfolioAsset
)

from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
from reportlab.lib import colors


def _sym(ctx, code: str) -> str:
    cur = ctx.currency.get_by_code(code)
    return cur.symbol if cur else ""


def _fmt(amount: float, symbol: str = "", decimals: int = 2) -> str:
    sign = "-" if amount < 0 else ""
    return f"{sign}{symbol}{abs(amount):,.{decimals}f}"


# ─────────────────────────────────────────────────────────────────────────────
# 1. NET WORTH REPORT
# ─────────────────────────────────────────────────────────────────────────────

class NetWorthReport:
    def generate(self, ctx, output_path: str) -> str:
        base = ctx.settings.get("base_currency", "USD")
        sym  = _sym(ctx, base)
        snap = ctx.account.net_worth_snapshot(base)
        port = ctx.portfolio.portfolio_summary(base)
        loan = ctx.loan.summary(base)

        rb = ReportBuilder("Net Worth Report", "Complete financial position", ctx, output_path)
        rb.add_cover([
            f"Base Currency: {base}",
            f"Accounts: {len(snap['accounts'])}",
            f"Portfolio Assets: {len(port['assets'])}",
        ])

        # Summary stats
        rb.section("Financial Position Overview")
        total_assets  = snap["total"] + port["total_value"]
        total_debts   = loan["i_owe"]
        true_net_worth = total_assets - total_debts + loan["owed_to_me"]

        rb.stat_row([
            ("Net Worth",       f"{sym}{true_net_worth:,.2f}", f"Total position in {base}", C_GOLD),
            ("Liquid Accounts", f"{sym}{snap['total']:,.2f}",  "Bank + wallet + savings",   C_ACCENT),
            ("Investments",     f"{sym}{port['total_value']:,.2f}", "Portfolio market value", C_GREEN),
            ("Owed to Me",      f"{sym}{loan['owed_to_me']:,.2f}", "Personal loans out",     C_GREEN),
        ])
        rb.stat_row([
            ("I Owe",           f"{sym}{loan['i_owe']:,.2f}",  "Personal debts",  C_RED),
            ("Portfolio P&L",   f"{sym}{port['total_pnl']:,.2f}", f"{port['pnl_pct']:+.2f}%", C_GREEN if port["total_pnl"] >= 0 else C_RED),
            ("Active Accounts", str(len(snap["accounts"])), "",  C_MUTED),
            ("Active Assets",   str(len(port["assets"])),   "",  C_MUTED),
        ])

        # Accounts breakdown table
        rb.section("Account Balances")
        headers = ["Account", "Type", "Institution", "Balance", "Currency", f"Value ({base})"]
        rows = []
        for acc in snap["accounts"]:
            rows.append([
                acc["name"],
                acc["type"],
                "",  # institution not in snap — load separately
                _money_cell(acc["balance"], acc.get("currency_symbol", "")),
                acc["currency"],
                _money_cell(acc["balance_base"], sym),
            ])
        rb.data_table(headers, rows, col_widths=[30, 22, 22, 22, 14, 22])

        # Pie chart – assets by account
        if snap["accounts"]:
            vals   = [max(a["balance_base"], 0) for a in snap["accounts"]]
            labels = [a["name"][:14] for a in snap["accounts"]]
            total  = sum(vals) or 1
            if any(v > 0 for v in vals):
                d = rb.pie_chart(vals, [f"{l}\n{v/total*100:.1f}%" for l, v in zip(labels, vals)],
                                 title="Asset Distribution by Account")
                rb.add_drawing(d)

        # Portfolio summary
        if port["assets"]:
            rb.section("Investment Portfolio Summary")
            headers = ["Ticker", "Name", "Type", "Qty", "Avg Cost", "Price", "Mkt Value", "P&L", "P&L %"]
            rows = []
            for a in port["assets"]:
                cur_sym = _sym(ctx, a["currency"])
                rows.append([
                    a["ticker"] or "—",
                    a["name"][:24],
                    a["type"],
                    f"{a['quantity']:,.4f}",
                    _money_cell(a["avg_cost"], cur_sym),
                    _money_cell(a["current_price"] or a["avg_cost"], cur_sym),
                    _money_cell(a["market_value_base"], sym),
                    _money_cell(a["unrealized_pnl"], cur_sym),
                    _pct_cell(a["pnl_pct"]),
                ])
            rb.data_table(headers, rows, col_widths=[14, 35, 16, 14, 18, 18, 22, 22, 14])

        # Loans
        loans = ctx.loan.get_all()
        if loans:
            rb.section("Personal Loans & Debts")
            headers = ["Contact", "Direction", "Principal", "Repaid", "Outstanding", "CCY", "Due Date"]
            rows = []
            for loan in loans:
                cur = loan.currency
                ls  = cur.symbol if cur else ""
                rows.append([
                    loan.contact_name,
                    "They owe me" if loan.direction == "owed_to_me" else "I owe",
                    _money_cell(loan.principal, ls),
                    _money_cell(loan.amount_repaid, ls),
                    _money_cell(loan.outstanding, ls),
                    cur.code if cur else "—",
                    loan.due_date.strftime("%d %b %Y") if loan.due_date else "—",
                ])
            rb.data_table(headers, rows, col_widths=[30, 20, 20, 20, 20, 12, 20])

        rb.divider()
        rb.paragraph("This report is generated from your local WealthMap database and is for personal use only.", "FootNote")

        return rb.save()


# ─────────────────────────────────────────────────────────────────────────────
# 2. ACCOUNT STATEMENT REPORT
# ─────────────────────────────────────────────────────────────────────────────

class AccountStatementReport:
    def generate(self, ctx, account_id: int, output_path: str,
                 date_from: Optional[datetime] = None,
                 date_to:   Optional[datetime] = None) -> str:
        from src.models.database import Account, Transaction

        acc = ctx.session.query(Account).get(account_id)
        if not acc:
            raise ValueError(f"Account {account_id} not found")

        base  = ctx.settings.get("base_currency", "USD")
        sym   = _sym(ctx, acc.currency.code)
        bsym  = _sym(ctx, base)

        # Date range
        if date_to is None:
            date_to = datetime.now(timezone.utc)
        if date_from is None:
            date_from = date_to.replace(day=1, hour=0, minute=0, second=0)

        txs = (ctx.session.query(Transaction)
               .filter(
                   Transaction.account_id == account_id,
                   Transaction.transaction_date >= date_from,
                   Transaction.transaction_date <= date_to,
               )
               .order_by(Transaction.transaction_date.asc())
               .all())

        bal = ctx.account.get_balance(acc)

        title = f"Account Statement — {acc.name}"
        subtitle = f"{date_from.strftime('%d %b %Y')} to {date_to.strftime('%d %b %Y')}"
        rb = ReportBuilder(title, subtitle, ctx, output_path)
        rb.add_cover([
            f"Account: {acc.name}",
            f"Type: {acc.account_type.value}",
            f"Institution: {acc.institution or 'N/A'}",
            f"Currency: {acc.currency.code}",
            f"Period: {subtitle}",
        ])

        # Summary
        rb.section("Account Summary")
        income   = sum(t.amount for t in txs if t.transaction_type in
                       (TransactionType.INCOME, TransactionType.DIVIDEND, TransactionType.LOAN_IN)
                       and t.status != TransactionStatus.CANCELLED)
        expenses = sum(t.amount for t in txs if t.transaction_type in
                       (TransactionType.EXPENSE, TransactionType.LOAN_OUT, TransactionType.INVESTMENT)
                       and t.status != TransactionStatus.CANCELLED)
        rb.stat_row([
            ("Current Balance", _fmt(bal, sym), acc.currency.code, C_GOLD),
            ("Period Income",   _fmt(income, sym),   f"{len(txs)} transactions", C_GREEN),
            ("Period Expenses", _fmt(expenses, sym), "",                          C_RED),
            ("Net Movement",    _fmt(income - expenses, sym), "",
             C_GREEN if income >= expenses else C_RED),
        ])

        # Transaction table
        rb.section("Transactions")
        headers = ["Date", "Type", "Description", "Category", "Payee", "Amount", "Status"]
        rows = []
        running = 0.0
        for tx in txs:
            is_credit = tx.transaction_type in (
                TransactionType.INCOME, TransactionType.DIVIDEND, TransactionType.LOAN_IN
            )
            amt = tx.amount if is_credit else -tx.amount
            running += amt
            rows.append([
                tx.transaction_date.strftime("%d %b %Y"),
                tx.transaction_type.value,
                (tx.description or "—")[:30],
                tx.category or "—",
                (tx.payee or "—")[:20],
                _money_cell(amt, sym),
                tx.status.value,
            ])
        rb.data_table(headers, rows, col_widths=[18, 20, 36, 22, 24, 20, 16])

        # Category breakdown
        by_cat: Dict[str, float] = {}
        for tx in txs:
            if tx.transaction_type in (TransactionType.EXPENSE, TransactionType.LOAN_OUT):
                by_cat[tx.category or "Other"] = by_cat.get(tx.category or "Other", 0) + tx.amount

        if by_cat:
            rb.section("Expense Breakdown by Category")
            cats   = sorted(by_cat.items(), key=lambda x: -x[1])
            total_exp = sum(v for _, v in cats) or 1
            headers2 = ["Category", "Amount", "% of Expenses"]
            rows2 = []
            for cat, amt in cats:
                rows2.append([
                    cat,
                    _money_cell(amt, sym),
                    _pct_cell(amt / total_exp * 100),
                ])
            rb.data_table(headers2, rows2, col_widths=[60, 30, 30])

            d = rb.pie_chart(
                [v for _, v in cats[:8]],
                [f"{k[:12]}\n{v/total_exp*100:.1f}%" for k, v in cats[:8]],
                title="Expenses by Category"
            )
            rb.add_drawing(d)

        rb.divider()
        rb.paragraph(f"Statement for account: {acc.name}  •  Account number: {acc.account_number or 'N/A'}", "FootNote")
        return rb.save()


# ─────────────────────────────────────────────────────────────────────────────
# 3. TRANSACTION HISTORY REPORT
# ─────────────────────────────────────────────────────────────────────────────

class TransactionReport:
    def generate(self, ctx, output_path: str,
                 year: int = None, month: int = None,
                 account_id: Optional[int] = None) -> str:
        from src.models.database import Transaction

        now   = datetime.now(timezone.utc)
        year  = year  or now.year
        month = month or now.month
        base  = ctx.settings.get("base_currency", "USD")
        bsym  = _sym(ctx, base)

        from sqlalchemy import extract
        q = ctx.session.query(Transaction).filter(
            extract("year",  Transaction.transaction_date) == year,
            extract("month", Transaction.transaction_date) == month,
        )
        if account_id:
            q = q.filter(Transaction.account_id == account_id)
        txs = q.order_by(Transaction.transaction_date.asc()).all()

        period = datetime(year, month, 1).strftime("%B %Y")
        rb = ReportBuilder(
            f"Transaction Report — {period}",
            f"{len(txs)} transactions",
            ctx, output_path
        )
        rb.add_cover([f"Period: {period}", f"Total transactions: {len(txs)}"])

        # Monthly summary
        monthly = ctx.transaction.monthly_summary(year, month, base)
        rb.section(f"Summary — {period}")
        rb.stat_row([
            ("Total Income",   _fmt(monthly["income"],   bsym), base, C_GREEN),
            ("Total Expenses", _fmt(monthly["expenses"], bsym), base, C_RED),
            ("Net Cash Flow",  _fmt(monthly["net"],      bsym), base,
             C_GREEN if monthly["net"] >= 0 else C_RED),
            ("Transactions",   str(len(txs)), f"in {period}", C_MUTED),
        ])

        # Category bar chart
        by_cat = monthly.get("by_category", {})
        if by_cat:
            cats_sorted = sorted(by_cat.items(), key=lambda x: -x[1])[:10]
            d = rb.bar_chart(
                [[v for _, v in cats_sorted]],
                [k[:12] for k, _ in cats_sorted],
                title=f"Expenses by Category — {period}",
                width=170, height=80
            )
            rb.add_drawing(d)

        # Full transaction table — page 2
        rb.page_break()
        rb.section("All Transactions")
        headers = ["Date", "Account", "Type", "Description", "Category", "Payee", "Amount", "CCY", "Status"]
        rows = []
        for tx in txs:
            is_credit = tx.transaction_type in (
                TransactionType.INCOME, TransactionType.DIVIDEND, TransactionType.LOAN_IN
            )
            amt = tx.amount if is_credit else -tx.amount
            cur = tx.currency
            s = cur.symbol if cur else ""
            rows.append([
                tx.transaction_date.strftime("%d %b"),
                (tx.account.name[:16] if tx.account else "—"),
                tx.transaction_type.value[:12],
                (tx.description or "—")[:28],
                (tx.category or "—")[:18],
                (tx.payee or "—")[:18],
                _money_cell(amt, s),
                (cur.code if cur else "—"),
                tx.status.value[:10],
            ])
        rb.data_table(headers, rows, col_widths=[14, 22, 18, 38, 22, 22, 18, 10, 14])

        return rb.save()


# ─────────────────────────────────────────────────────────────────────────────
# 4. PORTFOLIO REPORT
# ─────────────────────────────────────────────────────────────────────────────

class PortfolioReport:
    def generate(self, ctx, output_path: str) -> str:
        from src.models.database import PortfolioAsset, AssetTrade

        base = ctx.settings.get("base_currency", "USD")
        bsym = _sym(ctx, base)
        port = ctx.portfolio.portfolio_summary(base)

        rb = ReportBuilder("Portfolio Report", "Investment holdings & performance", ctx, output_path)
        rb.add_cover([
            f"Assets tracked: {len(port['assets'])}",
            f"Total market value: {bsym}{port['total_value']:,.2f} {base}",
            f"Unrealized P&L: {bsym}{port['total_pnl']:+,.2f}",
        ])

        # Summary
        rb.section("Portfolio Overview")
        rb.stat_row([
            ("Market Value",   f"{bsym}{port['total_value']:,.2f}", base, C_GOLD),
            ("Cost Basis",     f"{bsym}{port['total_cost']:,.2f}",  base, C_MUTED),
            ("Unrealized P&L", f"{bsym}{port['total_pnl']:+,.2f}",
             f"{port['pnl_pct']:+.2f}%", C_GREEN if port["total_pnl"] >= 0 else C_RED),
            ("Assets",         str(len(port["assets"])), "holdings", C_ACCENT),
        ])

        # Allocation pie by type
        by_type: Dict[str, float] = {}
        for a in port["assets"]:
            by_type[a["type"]] = by_type.get(a["type"], 0) + a["market_value_base"]

        if by_type:
            total_val = sum(by_type.values()) or 1
            d = rb.pie_chart(
                list(by_type.values()),
                [f"{k}\n{v/total_val*100:.1f}%" for k, v in by_type.items()],
                title="Allocation by Asset Type"
            )
            rb.add_drawing(d)

        # Holdings table
        rb.section("Holdings Detail")
        headers = ["Ticker", "Name", "Type", "Qty", "Avg Cost", "Curr Price",
                   f"Mkt Val ({base})", "P&L", "P&L %", "CCY"]
        rows = []
        for a in port["assets"]:
            cur_sym = _sym(ctx, a["currency"])
            rows.append([
                a["ticker"] or "—",
                a["name"][:22],
                a["type"][:12],
                f"{a['quantity']:,.4f}",
                _money_cell(a["avg_cost"], cur_sym),
                _money_cell(a["current_price"] or a["avg_cost"], cur_sym),
                _money_cell(a["market_value_base"], bsym),
                _money_cell(a["unrealized_pnl"], cur_sym),
                _pct_cell(a["pnl_pct"]),
                a["currency"],
            ])
        rb.data_table(headers, rows, col_widths=[12, 30, 14, 14, 18, 18, 22, 20, 14, 10])

        # Trade history
        rb.page_break()
        rb.section("Trade History")
        assets = ctx.session.query(PortfolioAsset).filter_by(is_active=True).all()
        for asset in assets:
            if not asset.trades:
                continue
            rb.sub_section(f"{asset.ticker or asset.name} — Trade Log")
            cur_sym = _sym(ctx, asset.currency.code)
            headers2 = ["Date", "Trade", "Qty", "Price", "Fees", "Total"]
            rows2 = []
            for trade in sorted(asset.trades, key=lambda t: t.trade_date):
                total = trade.quantity * trade.price + trade.fees
                rows2.append([
                    trade.trade_date.strftime("%d %b %Y"),
                    trade.trade_type,
                    f"{trade.quantity:,.4f}",
                    _money_cell(trade.price, cur_sym),
                    _money_cell(trade.fees, cur_sym),
                    _money_cell(total, cur_sym),
                ])
            rb.data_table(headers2, rows2, col_widths=[20, 14, 18, 22, 18, 22])

        return rb.save()


# ─────────────────────────────────────────────────────────────────────────────
# 5. LOANS REPORT
# ─────────────────────────────────────────────────────────────────────────────

class LoansReport:
    def generate(self, ctx, output_path: str, include_settled: bool = False) -> str:
        base = ctx.settings.get("base_currency", "USD")
        bsym = _sym(ctx, base)
        summary = ctx.loan.summary(base)
        loans = ctx.loan.get_all(include_settled=include_settled)

        rb = ReportBuilder("Loans & Personal Debts",
                           "Inter-personal transaction register", ctx, output_path)
        rb.add_cover([
            f"Active loans: {len(loans)}",
            f"Owed to me: {bsym}{summary['owed_to_me']:,.2f} {base}",
            f"I owe: {bsym}{summary['i_owe']:,.2f} {base}",
        ])

        rb.section("Summary")
        net_col = C_GREEN if summary["net"] >= 0 else C_RED
        rb.stat_row([
            ("Owed to Me",   f"{bsym}{summary['owed_to_me']:,.2f}", base, C_GREEN),
            ("I Owe",        f"{bsym}{summary['i_owe']:,.2f}",      base, C_RED),
            ("Net Position", f"{bsym}{summary['net']:,.2f}",        base, net_col),
            ("Active Loans", str(len(loans)), "records", C_MUTED),
        ])

        # Loans table
        rb.section("Loan Register")
        headers = ["Contact", "Direction", "Currency", "Principal",
                   "Repaid", "Outstanding", "% Done", "Due Date"]
        rows = []
        for loan in loans:
            cur = loan.currency
            ls  = cur.symbol if cur else ""
            pct = (loan.amount_repaid / loan.principal * 100) if loan.principal else 0
            rows.append([
                loan.contact_name,
                "They owe me" if loan.direction == "owed_to_me" else "I owe",
                cur.code if cur else "—",
                _money_cell(loan.principal, ls),
                _money_cell(loan.amount_repaid, ls),
                _money_cell(loan.outstanding, ls),
                _pct_cell(pct),
                loan.due_date.strftime("%d %b %Y") if loan.due_date else "—",
            ])
        rb.data_table(headers, rows, col_widths=[28, 20, 12, 20, 20, 20, 14, 18])

        # Per-loan repayment detail
        rb.page_break()
        rb.section("Repayment History (per loan)")
        for loan in loans:
            if not loan.repayments:
                continue
            cur = loan.currency
            ls  = cur.symbol if cur else ""
            rb.sub_section(f"{loan.contact_name}  —  {cur.code if cur else ''} {_fmt(loan.principal, ls)}")
            headers2 = ["Date", "Amount", "Notes"]
            rows2 = []
            for rep in sorted(loan.repayments, key=lambda r: r.repaid_on):
                rows2.append([
                    rep.repaid_on.strftime("%d %b %Y"),
                    _money_cell(rep.amount, ls),
                    rep.notes or "—",
                ])
            rb.data_table(headers2, rows2, col_widths=[28, 28, 100])

        return rb.save()


# ─────────────────────────────────────────────────────────────────────────────
# 6. RECEIPTS REPORT
# ─────────────────────────────────────────────────────────────────────────────

class ReceiptsReport:
    def generate(self, ctx, output_path: str) -> str:
        from src.models.database import Receipt

        receipts = (ctx.session.query(Receipt)
                    .order_by(Receipt.receipt_date.desc()).all())

        rb = ReportBuilder("Receipts Register",
                           f"{len(receipts)} receipts on file", ctx, output_path)
        rb.add_cover([f"Total receipts stored: {len(receipts)}"])

        rb.section("Receipts Index")
        headers = ["Date", "Title", "Merchant", "Amount", "CCY", "Category", "Files"]
        rows = []
        for r in receipts:
            cur = r.currency
            ls  = cur.symbol if cur else ""
            rows.append([
                r.receipt_date.strftime("%d %b %Y") if r.receipt_date else "—",
                (r.title or "—")[:34],
                (r.merchant or "—")[:22],
                _money_cell(r.amount, ls) if r.amount else Paragraph("—", STYLES["TableCell"]),
                cur.code if cur else "—",
                (r.category or "—")[:18],
                str(len(r.attachments)),
            ])
        rb.data_table(headers, rows, col_widths=[18, 46, 28, 22, 10, 22, 10])

        # Attachments index
        rb.section("Attached Files Index")
        all_atts = [(r, att) for r in receipts for att in r.attachments]
        if all_atts:
            headers2 = ["Receipt", "Filename", "Type", "Size (KB)", "Date Attached"]
            rows2 = []
            for r, att in all_atts:
                rows2.append([
                    (r.title or "—")[:28],
                    att.original_filename[:36],
                    att.mime_type.split("/")[-1].upper()[:10],
                    f"{att.file_size // 1024}",
                    att.created_at.strftime("%d %b %Y"),
                ])
            rb.data_table(headers2, rows2, col_widths=[38, 50, 14, 14, 22])
        else:
            rb.paragraph("No files attached to receipts yet.", "BodyMuted")

        return rb.save()


# ─────────────────────────────────────────────────────────────────────────────
# 7. EXCHANGE RATES REPORT
# ─────────────────────────────────────────────────────────────────────────────

class ExchangeRatesReport:
    def generate(self, ctx, output_path: str) -> str:
        from src.models.database import ExchangeRate

        rates = (ctx.session.query(ExchangeRate)
                 .order_by(ExchangeRate.base_currency_id,
                           ExchangeRate.target_currency_id).all())

        rb = ReportBuilder("Exchange Rates Report",
                           f"{len(rates)} rates on file", ctx, output_path)
        rb.add_cover([f"Cached rates: {len(rates)}"])

        rb.section("All Exchange Rates")
        headers = ["Base", "Target", "Rate", "Source", "Last Updated", "Age (min)"]
        rows = []
        for r in rates:
            age_min = int(
                (datetime.now(timezone.utc) -
                 r.fetched_at.replace(tzinfo=timezone.utc)).total_seconds() // 60
            )
            rows.append([
                r.base_currency.code if r.base_currency else "—",
                r.target_currency.code if r.target_currency else "—",
                f"{r.rate:,.6f}",
                r.source.upper(),
                r.fetched_at.strftime("%d %b %Y %H:%M"),
                str(age_min),
            ])
        rb.data_table(headers, rows, col_widths=[18, 18, 28, 14, 36, 18])

        # Key rates snapshot (base = USD)
        usd = ctx.currency.get_by_code("USD")
        if usd:
            major = ["EUR", "GBP", "JPY", "CHF", "CAD", "AUD", "CNY", "INR", "BTC", "ETH"]
            rb.section("Major Rates vs USD")
            headers2 = ["Currency", "Name", "Rate (1 USD =)", "Source"]
            rows2 = []
            for code in major:
                target = ctx.currency.get_by_code(code)
                if not target:
                    continue
                rate_row = (ctx.session.query(ExchangeRate)
                            .filter_by(base_currency_id=usd.id,
                                       target_currency_id=target.id)
                            .order_by(ExchangeRate.fetched_at.desc()).first())
                if rate_row:
                    rows2.append([
                        target.code,
                        target.name,
                        f"{rate_row.rate:,.6f} {target.symbol}",
                        rate_row.source.upper(),
                    ])
            if rows2:
                rb.data_table(headers2, rows2, col_widths=[18, 50, 40, 20])

        return rb.save()


# ─────────────────────────────────────────────────────────────────────────────
# 8. ANNUAL / MASTER FINANCIAL REPORT
# ─────────────────────────────────────────────────────────────────────────────

class AnnualReport:
    def generate(self, ctx, output_path: str, year: int = None) -> str:
        from sqlalchemy import extract
        from src.models.database import Transaction

        now  = datetime.now(timezone.utc)
        year = year or now.year
        base = ctx.settings.get("base_currency", "USD")
        bsym = _sym(ctx, base)

        rb = ReportBuilder(
            f"Annual Financial Report — {year}",
            "Comprehensive personal finance overview",
            ctx, output_path
        )

        # ── Cover ──────────────────────────────────────────────────────────────
        snap = ctx.account.net_worth_snapshot(base)
        port = ctx.portfolio.portfolio_summary(base)
        loan = ctx.loan.summary(base)
        rb.add_cover([
            f"Financial year: {year}",
            f"Base currency: {base}",
            f"Accounts: {len(snap['accounts'])}  |  "
            f"Portfolio assets: {len(port['assets'])}  |  "
            f"Active loans: {len(ctx.loan.get_all())}",
        ])

        # ── Section 1: Net Worth ───────────────────────────────────────────────
        rb.section("1. Net Worth", "Your total financial position")
        total_assets  = snap["total"] + port["total_value"]
        total_debts   = loan["i_owe"]
        net_worth     = total_assets - total_debts + loan["owed_to_me"]
        rb.stat_row([
            ("Net Worth",      f"{bsym}{net_worth:,.2f}",          base, C_GOLD),
            ("Liquid Assets",  f"{bsym}{snap['total']:,.2f}",      base, C_ACCENT),
            ("Investments",    f"{bsym}{port['total_value']:,.2f}", base, C_GREEN),
            ("Liabilities",    f"{bsym}{total_debts:,.2f}",        base, C_RED),
        ])

        # Account list
        headers = ["Account", "Type", "Balance", "CCY", f"{base} Value"]
        rows = []
        for acc in snap["accounts"]:
            rows.append([
                acc["name"], acc["type"],
                _money_cell(acc["balance"], ""),
                acc["currency"],
                _money_cell(acc["balance_base"], bsym),
            ])
        rb.data_table(headers, rows, col_widths=[38, 28, 24, 12, 24])

        # ── Section 2: Annual Income & Expenses ───────────────────────────────
        rb.page_break()
        rb.section("2. Annual Cash Flow", f"All transactions in {year}")

        monthly_data  = []
        income_series = []
        expense_series= []
        month_labels  = []

        for m in range(1, 13):
            ms = ctx.transaction.monthly_summary(year, m, base)
            monthly_data.append(ms)
            income_series.append(ms["income"])
            expense_series.append(ms["expenses"])
            month_labels.append(datetime(year, m, 1).strftime("%b"))

        total_income   = sum(d["income"]   for d in monthly_data)
        total_expenses = sum(d["expenses"] for d in monthly_data)
        total_net      = total_income - total_expenses

        rb.stat_row([
            ("Annual Income",   f"{bsym}{total_income:,.2f}",   base, C_GREEN),
            ("Annual Expenses", f"{bsym}{total_expenses:,.2f}", base, C_RED),
            ("Annual Net",      f"{bsym}{total_net:,.2f}",      base,
             C_GREEN if total_net >= 0 else C_RED),
            ("Savings Rate",
             f"{total_net / total_income * 100:.1f}%" if total_income else "—",
             "", C_ACCENT),
        ])

        # Monthly income vs expenses bar chart
        d = rb.bar_chart(
            [income_series, expense_series],
            month_labels,
            series_names=["Income", "Expenses"],
            title=f"Monthly Income vs Expenses — {year}",
            width=170, height=85
        )
        rb.add_drawing(d)

        # Monthly table
        rb.sub_section("Monthly Breakdown")
        headers2 = ["Month", "Income", "Expenses", "Net", "Savings Rate"]
        rows2 = []
        for m, ms in enumerate(monthly_data, 1):
            sr = (ms["net"] / ms["income"] * 100) if ms["income"] else 0
            rows2.append([
                datetime(year, m, 1).strftime("%B"),
                _money_cell(ms["income"],   bsym),
                _money_cell(ms["expenses"], bsym),
                _money_cell(ms["net"],      bsym),
                _pct_cell(sr),
            ])
        rb.data_table(headers2, rows2, col_widths=[28, 28, 28, 28, 24])

        # ── Section 3: Spending by Category ───────────────────────────────────
        rb.page_break()
        rb.section("3. Spending Analysis")
        all_by_cat: Dict[str, float] = {}
        for ms in monthly_data:
            for cat, amt in ms.get("by_category", {}).items():
                all_by_cat[cat] = all_by_cat.get(cat, 0) + amt

        if all_by_cat:
            cats_sorted = sorted(all_by_cat.items(), key=lambda x: -x[1])
            total_sp    = sum(v for _, v in cats_sorted) or 1
            headers3 = ["Category", "Annual Spend", "% of Total", "Avg Monthly"]
            rows3 = []
            for cat, amt in cats_sorted:
                rows3.append([
                    cat,
                    _money_cell(amt, bsym),
                    _pct_cell(amt / total_sp * 100),
                    _money_cell(amt / 12, bsym),
                ])
            rb.data_table(headers3, rows3, col_widths=[48, 28, 24, 28])

            # Pie chart – top 8 categories
            d2 = rb.pie_chart(
                [v for _, v in cats_sorted[:8]],
                [f"{k[:12]}\n{v/total_sp*100:.1f}%" for k, v in cats_sorted[:8]],
                title=f"Spending by Category — {year}"
            )
            rb.add_drawing(d2)

        # ── Section 4: Portfolio ───────────────────────────────────────────────
        if port["assets"]:
            rb.page_break()
            rb.section("4. Investment Portfolio")
            rb.stat_row([
                ("Market Value",   f"{bsym}{port['total_value']:,.2f}", base, C_GOLD),
                ("Unrealized P&L", f"{bsym}{port['total_pnl']:+,.2f}",
                 f"{port['pnl_pct']:+.2f}%", C_GREEN if port["total_pnl"] >= 0 else C_RED),
                ("Assets",         str(len(port["assets"])), "holdings", C_ACCENT),
                ("Cost Basis",     f"{bsym}{port['total_cost']:,.2f}", base, C_MUTED),
            ])
            headers4 = ["Ticker", "Name", "Qty", "Avg Cost", "Price",
                        f"Value ({base})", "P&L", "P&L %"]
            rows4 = []
            for a in port["assets"]:
                cs = _sym(ctx, a["currency"])
                rows4.append([
                    a["ticker"] or "—", a["name"][:22],
                    f"{a['quantity']:,.4f}",
                    _money_cell(a["avg_cost"], cs),
                    _money_cell(a["current_price"] or a["avg_cost"], cs),
                    _money_cell(a["market_value_base"], bsym),
                    _money_cell(a["unrealized_pnl"], cs),
                    _pct_cell(a["pnl_pct"]),
                ])
            rb.data_table(headers4, rows4, col_widths=[12, 30, 14, 18, 18, 22, 20, 14])

        # ── Section 5: Loans summary ───────────────────────────────────────────
        loans = ctx.loan.get_all()
        if loans:
            rb.page_break()
            rb.section("5. Personal Loans & Debts")
            rb.stat_row([
                ("Owed to Me",   f"{bsym}{loan['owed_to_me']:,.2f}", base, C_GREEN),
                ("I Owe",        f"{bsym}{loan['i_owe']:,.2f}",      base, C_RED),
                ("Net",          f"{bsym}{loan['net']:,.2f}",        base,
                 C_GREEN if loan["net"] >= 0 else C_RED),
                ("Open Loans",   str(len(loans)), "active", C_MUTED),
            ])
            headers5 = ["Contact", "Direction", "CCY", "Principal", "Outstanding"]
            rows5 = []
            for ln in loans:
                cur = ln.currency
                ls  = cur.symbol if cur else ""
                rows5.append([
                    ln.contact_name,
                    "They owe me" if ln.direction == "owed_to_me" else "I owe",
                    cur.code if cur else "—",
                    _money_cell(ln.principal, ls),
                    _money_cell(ln.outstanding, ls),
                ])
            rb.data_table(headers5, rows5, col_widths=[38, 24, 12, 24, 24])

        # Footer / disclaimer
        rb.page_break()
        rb.section("Notes & Disclaimer")
        rb.paragraph(
            "This report is generated automatically from your local WealthMap database. "
            "All values are for informational purposes only and should not be used as "
            "formal financial statements. Exchange rates used are cached rates and may "
            "not reflect exact market rates at the time of each transaction. "
            "For tax and accounting purposes, please consult a qualified professional.", "Body"
        )
        rb.divider()
        rb.paragraph(f"WealthMap Annual Report {year}  •  Base Currency: {base}  •  Local Edition", "FootNote")

        return rb.save()


# ─────────────────────────────────────────────────────────────────────────────
# 9. XLSX EXPORT (all-in-one spreadsheet)
# ─────────────────────────────────────────────────────────────────────────────

class XLSXExporter:
    """Generates a multi-sheet Excel workbook with all financial data."""

    def generate(self, ctx, output_path: str) -> str:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter

        base = ctx.settings.get("base_currency", "USD")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Colours
        HDR_FILL  = PatternFill("solid", fgColor="161B22")
        ROW_FILL1 = PatternFill("solid", fgColor="0D1117")
        ROW_FILL2 = PatternFill("solid", fgColor="1A2130")
        ACC_FONT  = Font(color="58A6FF", bold=True, name="Arial", size=9)
        HDR_FONT  = Font(color="FFFFFF", bold=True, name="Arial", size=9)
        BODY_FONT = Font(color="E6EDF3", name="Arial", size=9)
        GRN_FONT  = Font(color="3FB950", bold=True, name="Arial", size=9)
        RED_FONT  = Font(color="F85149", bold=True, name="Arial", size=9)
        GOLD_FONT = Font(color="E3B341", bold=True, name="Arial", size=10)
        CENTER    = Alignment(horizontal="center", vertical="center")
        RIGHT     = Alignment(horizontal="right")
        LEFT      = Alignment(horizontal="left")
        thin      = Side(style="thin", color="30363D")
        BORDER    = Border(bottom=thin)

        def hdr_row(ws, row_num, values, col_widths=None):
            for i, v in enumerate(values, 1):
                c = ws.cell(row=row_num, column=i, value=v)
                c.font = HDR_FONT
                c.fill = HDR_FILL
                c.alignment = CENTER
                c.border = Border(bottom=Side(style="medium", color="58A6FF"))
            if col_widths:
                for i, w in enumerate(col_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = w

        def body_row(ws, row_num, values, fills=None, fonts=None, aligns=None):
            fill = ROW_FILL1 if row_num % 2 == 0 else ROW_FILL2
            for i, v in enumerate(values, 1):
                c = ws.cell(row=row_num, column=i, value=v)
                c.font  = (fonts[i-1]  if fonts  and i <= len(fonts)  else BODY_FONT)
                c.fill  = (fills[i-1]  if fills  and i <= len(fills)  else fill)
                c.alignment = (aligns[i-1] if aligns and i <= len(aligns) else LEFT)
                c.border = Border(bottom=thin)

        # ── Sheet 1: Net Worth Summary ─────────────────────────────────────────
        ws = wb.create_sheet("Net Worth")
        snap = ctx.account.net_worth_snapshot(base)
        port = ctx.portfolio.portfolio_summary(base)
        loan = ctx.loan.summary(base)

        ws["A1"] = "WEALTHMAP — NET WORTH SUMMARY"
        ws["A1"].font = Font(color="58A6FF", bold=True, size=14, name="Arial")
        ws["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}   Base: {base}"
        ws["A2"].font = Font(color="8B949E", size=8, name="Arial")
        ws.merge_cells("A1:F1")

        row = 4
        hdr_row(ws, row, ["Account", "Type", "Institution", "Balance", "Currency", f"Value ({base})"],
                col_widths=[28, 20, 22, 16, 10, 18])
        row += 1
        for acc in snap["accounts"]:
            bal_font = GRN_FONT if acc["balance"] >= 0 else RED_FONT
            body_row(ws, row, [
                acc["name"], acc["type"], "",
                acc["balance"], acc["currency"], acc["balance_base"]
            ], fonts=[BODY_FONT, BODY_FONT, BODY_FONT, bal_font, BODY_FONT, GOLD_FONT],
               aligns=[LEFT, LEFT, LEFT, RIGHT, CENTER, RIGHT])
            row += 1

        row += 1
        ws.cell(row=row, column=5, value="TOTAL").font = HDR_FONT
        ws.cell(row=row, column=6, value=f"=SUM(F5:F{row-2})").font = GOLD_FONT
        ws.cell(row=row, column=6).alignment = RIGHT

        # ── Sheet 2: Transactions ──────────────────────────────────────────────
        ws2 = wb.create_sheet("Transactions")
        txs = ctx.transaction.get_recent(5000)
        hdr_row(ws2, 1, ["Date", "Account", "Type", "Description", "Category",
                          "Payee", "Reference", "Amount", "Currency", "Exchange Rate",
                          "Base Amount", "Status", "Notes"],
                col_widths=[12, 20, 16, 30, 18, 20, 18, 12, 8, 12, 12, 12, 24])
        for i, tx in enumerate(txs, 2):
            is_credit = tx.transaction_type in (
                TransactionType.INCOME, TransactionType.DIVIDEND, TransactionType.LOAN_IN
            )
            amt = tx.amount if is_credit else -tx.amount
            cur = tx.currency
            amt_font = GRN_FONT if is_credit else RED_FONT
            body_row(ws2, i, [
                tx.transaction_date.strftime("%Y-%m-%d"),
                tx.account.name if tx.account else "",
                tx.transaction_type.value,
                tx.description or "",
                tx.category or "",
                tx.payee or "",
                tx.reference or "",
                amt,
                cur.code if cur else "",
                tx.exchange_rate or "",
                tx.base_amount or "",
                tx.status.value,
                tx.notes or "",
            ], fonts=[BODY_FONT]*6 + [BODY_FONT, amt_font] + [BODY_FONT]*5,
               aligns=[LEFT]*7 + [RIGHT, CENTER, RIGHT, RIGHT] + [LEFT]*2)

        # ── Sheet 3: Portfolio ─────────────────────────────────────────────────
        ws3 = wb.create_sheet("Portfolio")
        hdr_row(ws3, 1, ["Ticker", "Name", "Type", "Qty", "Avg Cost", "Curr Price",
                          f"Mkt Value ({base})", "P&L", "P&L %", "Currency"],
                col_widths=[10, 28, 14, 12, 14, 14, 18, 16, 10, 10])
        for i, a in enumerate(port["assets"], 2):
            pnl_font = GRN_FONT if a["unrealized_pnl"] >= 0 else RED_FONT
            body_row(ws3, i, [
                a["ticker"] or "",
                a["name"],
                a["type"],
                a["quantity"],
                a["avg_cost"],
                a["current_price"] or a["avg_cost"],
                a["market_value_base"],
                a["unrealized_pnl"],
                a["pnl_pct"] / 100,
                a["currency"],
            ], fonts=[BODY_FONT]*6 + [GOLD_FONT, pnl_font, pnl_font, BODY_FONT],
               aligns=[CENTER, LEFT, CENTER, RIGHT, RIGHT, RIGHT, RIGHT, RIGHT, RIGHT, CENTER])
            ws3.cell(row=i, column=9).number_format = "0.00%"

        # ── Sheet 4: Loans ─────────────────────────────────────────────────────
        ws4 = wb.create_sheet("Loans")
        loans = ctx.loan.get_all(include_settled=True)
        hdr_row(ws4, 1, ["Contact", "Contact Info", "Direction", "Currency",
                          "Principal", "Repaid", "Outstanding", "Settled?", "Due Date"],
                col_widths=[22, 28, 16, 10, 14, 14, 14, 10, 14])
        for i, ln in enumerate(loans, 2):
            cur = ln.currency
            out_font = GRN_FONT if ln.direction == "owed_to_me" else RED_FONT
            body_row(ws4, i, [
                ln.contact_name,
                ln.contact_info or "",
                "They owe me" if ln.direction == "owed_to_me" else "I owe",
                cur.code if cur else "",
                ln.principal,
                ln.amount_repaid,
                ln.outstanding,
                "Yes" if ln.is_settled else "No",
                ln.due_date.strftime("%Y-%m-%d") if ln.due_date else "",
            ], fonts=[BODY_FONT]*4 + [BODY_FONT, BODY_FONT, out_font, BODY_FONT, BODY_FONT],
               aligns=[LEFT]*3 + [CENTER] + [RIGHT]*3 + [CENTER, CENTER])

        # ── Sheet 5: Monthly Summary ───────────────────────────────────────────
        ws5 = wb.create_sheet("Monthly Summary")
        year = datetime.now().year
        hdr_row(ws5, 1, ["Year", "Month", "Income", "Expenses", "Net", "Savings Rate"],
                col_widths=[8, 12, 16, 16, 16, 14])
        row = 2
        for yr in range(year - 1, year + 1):
            for m in range(1, 13):
                ms = ctx.transaction.monthly_summary(yr, m, base)
                if ms["income"] == 0 and ms["expenses"] == 0:
                    continue
                sr = ms["net"] / ms["income"] if ms["income"] else 0
                net_font = GRN_FONT if ms["net"] >= 0 else RED_FONT
                body_row(ws5, row, [
                    yr,
                    datetime(yr, m, 1).strftime("%B"),
                    ms["income"], ms["expenses"], ms["net"], sr
                ], fonts=[BODY_FONT, BODY_FONT, GRN_FONT, RED_FONT, net_font, BODY_FONT],
                   aligns=[CENTER, LEFT, RIGHT, RIGHT, RIGHT, RIGHT])
                ws5.cell(row=row, column=6).number_format = "0.00%"
                row += 1

        # ── Sheet 6: Exchange Rates ────────────────────────────────────────────
        ws6 = wb.create_sheet("Exchange Rates")
        from src.models.database import ExchangeRate
        rates = ctx.session.query(ExchangeRate).all()
        hdr_row(ws6, 1, ["Base", "Target", "Rate", "Source", "Updated"],
                col_widths=[10, 10, 16, 12, 22])
        for i, r in enumerate(rates, 2):
            body_row(ws6, i, [
                r.base_currency.code if r.base_currency else "",
                r.target_currency.code if r.target_currency else "",
                r.rate,
                r.source,
                r.fetched_at.strftime("%Y-%m-%d %H:%M"),
            ], aligns=[CENTER, CENTER, RIGHT, CENTER, LEFT])

        wb.save(output_path)
        return output_path


# ─────────────────────────────────────────────────────────────────────────────
# 11. FEES & TAXES REPORT
# ─────────────────────────────────────────────────────────────────────────────

class FeesTaxesReport:
    """
    Surfaces every fee and tax charged across transactions, personal loans
    (and their repayments), and portfolio trades — so the user can see
    exactly how much "extra" they're paying on top of their main expenses.
    """

    def generate(self, ctx, output_path: str) -> str:
        base = ctx.settings.get("base_currency", "USD")
        bsym = _sym(ctx, base)

        rb = ReportBuilder("Fees & Taxes Report",
                           "Every additional charge across transactions, loans & trades",
                           ctx, output_path)

        # ── Gather data ──────────────────────────────────────────────────
        txs = ctx.session.query(Transaction).all()

        tx_fee_total = tx_tax_total = 0.0
        tx_rows = []
        charge_rows = []
        by_category: Dict[str, float] = {}
        for tx in txs:
            fee = tx.fee_amount or 0.0
            tax = tx.tax_amount or 0.0
            if fee == 0 and tax == 0 and not tx.charges:
                continue
            cur = tx.currency
            sym = cur.symbol if cur else ""
            fee_base = ctx.currency.convert(fee, cur.code, base) or fee if fee else 0.0
            tax_base = ctx.currency.convert(tax, cur.code, base) or tax if tax else 0.0
            tx_fee_total += fee_base
            tx_tax_total += tax_base
            by_category[tx.category] = by_category.get(tx.category, 0) + fee_base + tax_base
            if fee or tax:
                tx_rows.append([
                    tx.transaction_date.strftime("%d %b %Y"),
                    tx.account.name if tx.account else "—",
                    tx.description or tx.category,
                    _money_cell(fee, sym) if fee else "—",
                    tx.fee_description or "—",
                    _money_cell(tax, sym) if tax else "—",
                    tx.tax_description or "—",
                ])
            for charge in tx.charges:
                c_cur = charge.currency or cur
                c_sym = c_cur.symbol if c_cur else sym
                c_base = ctx.currency.convert(charge.amount, c_cur.code, base) or charge.amount
                if charge.kind == "tax":
                    tx_tax_total += c_base
                else:
                    tx_fee_total += c_base
                by_category[tx.category] = by_category.get(tx.category, 0) + c_base
                charge_rows.append([
                    tx.transaction_date.strftime("%d %b %Y"),
                    tx.account.name if tx.account else "—",
                    tx.description or tx.category,
                    "Tax" if charge.kind == "tax" else "Fee",
                    _money_cell(charge.amount, c_sym),
                    charge.description or "—",
                ])

        # Loan fees
        loans = ctx.loan.get_all(include_settled=True)
        loan_fee_total = 0.0
        loan_rows = []
        repayment_fee_total = 0.0
        for loan in loans:
            fee = loan.fee_amount or 0.0
            cur = loan.currency
            sym = cur.symbol if cur else ""
            fee_base = (ctx.currency.convert(fee, cur.code, base) or fee) if fee else 0.0
            loan_fee_total += fee_base
            for rep in loan.repayments:
                rfee = rep.fee_amount or 0.0
                rfee_base = (ctx.currency.convert(rfee, cur.code, base) or rfee) if rfee else 0.0
                repayment_fee_total += rfee_base
            if fee == 0 and not any((r.fee_amount or 0) for r in loan.repayments):
                continue
            loan_rows.append([
                loan.contact_name,
                "They owe me" if loan.direction == "owed_to_me" else "I owe",
                _money_cell(fee, sym) if fee else "—",
                loan.fee_description or "—",
                _money_cell(sum(r.fee_amount or 0 for r in loan.repayments), sym),
            ])

        # Portfolio trade fees & taxes
        assets = ctx.session.query(PortfolioAsset).all()
        trade_fee_total = trade_tax_total = 0.0
        trade_rows = []
        for asset in assets:
            cur = asset.currency
            sym = cur.symbol if cur else ""
            for trade in asset.trades:
                fee = trade.fees or 0.0
                tax = trade.taxes or 0.0
                if fee == 0 and tax == 0:
                    continue
                fee_base = (ctx.currency.convert(fee, cur.code, base) or fee) if fee else 0.0
                tax_base = (ctx.currency.convert(tax, cur.code, base) or tax) if tax else 0.0
                trade_fee_total += fee_base
                trade_tax_total += tax_base
                trade_rows.append([
                    trade.trade_date.strftime("%d %b %Y"),
                    f"{asset.name} ({asset.ticker})" if asset.ticker else asset.name,
                    trade.trade_type,
                    _money_cell(fee, sym) if fee else "—",
                    _money_cell(tax, sym) if tax else "—",
                ])

        grand_total = tx_fee_total + tx_tax_total + loan_fee_total + repayment_fee_total + \
                       trade_fee_total + trade_tax_total

        # ── Cover & summary ─────────────────────────────────────────────
        rb.add_cover([
            f"Total fees & taxes paid (in {base}): {bsym}{grand_total:,.2f}",
            f"Transaction fees+taxes: {bsym}{(tx_fee_total + tx_tax_total):,.2f}",
            f"Loan-related fees: {bsym}{(loan_fee_total + repayment_fee_total):,.2f}",
            f"Portfolio trade fees+taxes: {bsym}{(trade_fee_total + trade_tax_total):,.2f}",
        ])

        rb.section("Summary")
        rb.stat_row([
            ("Transaction Fees", f"{bsym}{tx_fee_total:,.2f}", base, C_ACCENT),
            ("Transaction Taxes", f"{bsym}{tx_tax_total:,.2f}", base, C_GOLD),
            ("Loan Fees", f"{bsym}{(loan_fee_total + repayment_fee_total):,.2f}", base, C_RED),
            ("Trade Fees + Taxes", f"{bsym}{(trade_fee_total + trade_tax_total):,.2f}", base, C_MUTED),
        ])

        if by_category:
            rb.section("Fees & Taxes by Category")
            cats = sorted(by_category.items(), key=lambda kv: kv[1], reverse=True)
            labels = [c for c, _ in cats][:8]
            values = [v for _, v in cats][:8]
            if any(v > 0 for v in values):
                d = rb.pie_chart(values, [f"{l}\n{bsym}{v:,.2f}" for l, v in zip(labels, values)],
                                 title=f"By Category ({base})")
                rb.add_drawing(d)

        # ── Transaction detail ──────────────────────────────────────────
        if tx_rows:
            rb.section("Transaction Fees & Taxes")
            headers = ["Date", "Account", "Transaction", "Fee", "Fee Note", "Tax", "Tax Note"]
            rb.data_table(headers, tx_rows, col_widths=[18, 20, 30, 14, 24, 14, 24])
        else:
            rb.section("Transaction Fees & Taxes")
            rb.paragraph("No transactions have recorded fees or taxes yet.", style="BodyMuted")

        # ── Additional charges detail ────────────────────────────────────
        if charge_rows:
            rb.section("Additional Fees & Taxes")
            headers = ["Date", "Account", "Transaction", "Kind", "Amount", "Description"]
            rb.data_table(headers, charge_rows, col_widths=[18, 20, 30, 12, 16, 28])

        # ── Loan detail ──────────────────────────────────────────────────
        if loan_rows:
            rb.section("Loan-Related Fees")
            headers = ["Contact", "Direction", "Origination Fee", "Fee Note", "Total Repayment Fees"]
            rb.data_table(headers, loan_rows, col_widths=[26, 18, 18, 24, 18])

        # ── Trade detail ───────────────────────────────────────────────
        if trade_rows:
            rb.section("Portfolio Trade Fees & Taxes")
            headers = ["Date", "Asset", "Trade Type", "Fees", "Taxes"]
            rb.data_table(headers, trade_rows, col_widths=[16, 30, 14, 14, 14])

        return rb.save()
